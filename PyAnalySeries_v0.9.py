#!/usr/bin/env python

#=========================================================================================
# Author: Patrick Brockmann CEA/DRF/LSCE - September 2024
#=========================================================================================

#=========================================================================================
import sys
import os
import re
import numpy as np
import pandas as pd

import matplotlib.pyplot as plt
from matplotlib.patches import ConnectionPatch
from matplotlib.lines import Line2D
import matplotlib.patches as patches
import matplotlib as mpl

from scipy import interpolate

from openpyxl.utils import get_column_letter

#=========================================================================================
mpl.rcParams['toolbar'] = 'None'
mpl.rcParams['axes.labelsize'] = 10
mpl.rcParams['xtick.labelsize'] = 8
mpl.rcParams['ytick.labelsize'] = 8

version = 'v0.9'
curve1Color = 'darkmagenta'
curve2Color = 'forestgreen'
pointerColor = 'blue'
curveWidth = 0.8 

#=========================================================================================
usage = f"""
####################################################################################################################
Usage:  PyAnalySeries_{version}.py [-h]
        [-k kindInterpolation]
        fileXLSX

Options:
        -h, -?, --help, -help
                Print this manual
        -k, --kind
                Interpolation kind 'linear' or 'quadratic' (default 'linear')

Examples:
        PyAnalySeries_{version}.py testFile.xlsx
"""

#=========================================================================================
interactions = """
####################################################################################################################
Interactions:

-------------------------------------------------------------------------------
Press 'h'
    Display this help 
-------------------------------------------------------------------------------
Hold shift key while right click on a curve
    Create or move a pointer
-------------------------------------------------------------------------------
Hold down ctrl key on a plot
    Display points of the curve
-------------------------------------------------------------------------------
Hold down ctrl key on a plot while right click on a curve
    Create or move a pointer hooked on a point
-------------------------------------------------------------------------------
Press 'c' key
    Connect pointers
-------------------------------------------------------------------------------
Hold down x key while right click on a connection
    Delete the connection and its associated pointers
-------------------------------------------------------------------------------
Use wheel mouse on a plot
    Zoom in/out in the plot
-------------------------------------------------------------------------------
Hold down left key mouse on a plot
    Pan in the plot
-------------------------------------------------------------------------------
Hold down left key mouse on a plot
    Expand horizontal/vertical axis depending horizontal/vertical movement
-------------------------------------------------------------------------------
Press 'a' key on a plot
    Plot the 2 curves with an automatic vertical range and a horizontal range according to pointers
-------------------------------------------------------------------------------
Press 'A' key on a plot
    Plot the curve with automatic vertical and horizontal ranges
-------------------------------------------------------------------------------
Press 'p' key
    Save figure as pdf file and png file
-------------------------------------------------------------------------------
Press 'X' key
    Delete all pointers and connections 
-------------------------------------------------------------------------------
Press 'z' key
    Display/Hide interpolated curve
-------------------------------------------------------------------------------
Press 's' key
    Save data and pointers as excel file
-------------------------------------------------------------------------------
Press 'q' key
    Quit the application
"""
#=========================================================================================
kindInterpolation = 'linear'

while len(sys.argv[1:]) != 0:
    if sys.argv[1] in ('-h', '--help'):
        del(sys.argv[1])
        print(usage)
        sys.exit(1)
    elif sys.argv[1] in ('-k', '--kind'):
        kindInterpolation = sys.argv[2]
        del(sys.argv[1])
        del(sys.argv[1])
    elif re.match('-', sys.argv[1]):
        print('Unknown option')
        break
    else:
        break

if len(sys.argv[1:]) != 1:
    print(usage)
    sys.exit(1)

# -------------------------
if kindInterpolation not in ('linear', 'quadratic'):
    print(usage)
    sys.exit(1)

# -------------------------
fileData = sys.argv[1]

#=========================================================================================
key_x = False
key_shift = False
key_control = False
vline1 = None
vline2 = None
press = None
press_origin = None
cur_xlim = None
cur_ylim = None
xpress = None
ypress = None
mousepress = None
artistsList_Dict = {} 
vline1List = []
vline2List = []
coordsX1 = []
coordsX2 = []
curve1 = None
curve2 = None
curve2Interp = None
x1 = None
y1 = None
x2 = None
y2 = None
x2Interp = None
showInterp = False
second_xaxis = None

#=========================================================================================
def loadData(fileName):
    global x1, y1, x2, y2, x1Name, y1Name, x2Name, y2Name
    global coordsX1, coordsX2

    dataframe = pd.read_excel(fileName)
    #x1Name = 'Time (ka)'
    #y1Name = 'Stack Benthic d18O (per mil)'
    #x2Name = 'depthODP849cm'
    #y2Name = 'd18Oforams-b'
    x1Name, y1Name, x2Name, y2Name = dataframe.columns[0:4]    # First 4 columns
    x1 = dataframe[x1Name].to_numpy()
    y1 = dataframe[y1Name].to_numpy()
    x2 = dataframe[x2Name].to_numpy()
    y2 = dataframe[y2Name].to_numpy()

    try:
        dataframe = pd.read_excel(fileName, sheet_name="Pointers")
        coordsX1 = dataframe["Coordinates X1"].to_numpy()
        coordsX2 = dataframe["Coordinates X2"].to_numpy()

        # check if arrays are monotonically increasing
        if not (((np.diff(coordsX1) >= 0).all()) and ((np.diff(coordsX2) >= 0).all())):
            print(dataframe.to_string(index=False, header=False, float_format="%.8f"))
            print("Error: pointer coordinates are not monotonically increasing")
            coordsX1 = []
            coordsX2 = []

    except:
        print("No sheetname Pointers found in ", fileName)

#=========================================================================================
def updateConnections():

    for artistsList in artistsList_Dict.values():
        if isinstance(artistsList[0], ConnectionPatch):
            connect = artistsList[0]
            x1, y1 = connect.xy1
            connect.xy1 = (x1, axs[0].get_ylim()[0]) 
            x2, y2 = connect.xy2
            connect.xy2 = (x2, axs[1].get_ylim()[1])
            if ((axs[0].get_xlim()[0] < x1 < axs[0].get_xlim()[1]) and
                (axs[1].get_xlim()[0] < x2 < axs[1].get_xlim()[1])):
                connect.set_visible(True)
            else:
                connect.set_visible(False)

#=========================================================================================
def deleteConnections():
    global artistsList_Dict, vline1List, vline2List, coordsX1, coordsX2

    for objectId in artistsList_Dict.keys():
        for artist in artistsList_Dict[objectId]:
            artist.remove()
    artistsList_Dict = {}
    vline1List = []
    vline2List = []

    coordsX1 = []
    coordsX2 = []

#=========================================================================================
def drawConnections():
    global artistsList_Dict, vline1List, vline2List, coordsX1, coordsX2

    for i in range(len(coordsX1)):
        coordX1 = coordsX1[i]
        coordX2 = coordsX2[i]
        vline1 = axs[0].axvline(coordX1, color=pointerColor, alpha=0.5, linestyle='--', linewidth=1, label='vline')
        vline2 = axs[1].axvline(coordX2, color=pointerColor, alpha=0.5, linestyle='--', linewidth=1, label='vline')
        vline1List.append(vline1)
        vline2List.append(vline2)
        connect = ConnectionPatch(color=pointerColor, alpha=0.5, linewidth=1, picker=5, clip_on=True, label='connection',
                    xyA=(coordX1, axs[0].get_ylim()[0]), coordsA=axs[0].transData,
                    xyB=(coordX2, axs[1].get_ylim()[1]), coordsB=axs[1].transData)
        fig.add_artist(connect)
        artistsList_Dict[id(connect)] = [connect, vline1, vline2]

    updateConnections()
    setInterp()

#=========================================================================================
def deleteInterp():
    global x2Interp, curve2Interp, second_xaxis, showInterp

    if curve2Interp:
        curve2Interp.remove()
        curve2Interp = None
        x2Interp = None
        second_xaxis.remove()
        second_xaxis = None

#=========================================================================================
def setInterp():
    global x2Interp, curve2Interp, second_xaxis, coordsX1, coordsX2, showInterp

    if len(vline1List) <= 1:
        print("Warning: interpolation needs a minimum of 2 connections")
        if showInterp:
            showInterp = False
            displayInterp(showInterp)
        return

    deleteInterp()
    coordsX1 = sorted([float(line.get_xdata()[0]) for line in vline1List])
    coordsX2 = sorted([float(line.get_xdata()[0]) for line in vline2List])

    f_1to2 = interpolate.interp1d(coordsX1, coordsX2, kind=kindInterpolation, fill_value="extrapolate")
    f_2to1 = interpolate.interp1d(coordsX2, coordsX1, kind=kindInterpolation, fill_value="extrapolate")
    second_xaxis = axsInterp.secondary_xaxis('top', functions=(f_1to2, f_2to1))
    second_xaxis.tick_params(labelrotation=30)
    second_xaxis.set_xlabel(x2Name, color=curve2Color)
    plt.setp(second_xaxis.get_xticklabels(), horizontalalignment='left')

    x2Interp = f_2to1(x2)
    curve2Interp, = axsInterp.plot(x2Interp, y2, color=curve2Color, alpha=0.8, linewidth=curveWidth)

#=========================================================================================
def displayInterp(visible):

    if visible:
            axsInterp.set_visible(True)
            if curve2Interp: curve2Interp.set_visible(True)
    else:
            axsInterp.set_visible(False)
            if curve2Interp: curve2Interp.set_visible(False)
    axsInterp.figure.canvas.draw()

#=========================================================================================
def onPick(event):
    global vline1, vline2, artistsList_Dict, vline1List, vline2List

    artistLabel = event.artist.get_label()
    #print(artistLabel)

    #-----------------------------------------------
    if artistLabel == 'connection':
        if key_x:
            objectId = id(event.artist)
            for artist in artistsList_Dict[objectId]:
                artist.remove()
                if artist in vline1List:
                    vline1List.remove(artist)
                if artist in vline2List:
                    vline2List.remove(artist)
            del artistsList_Dict[objectId]
            setInterp()
            displayInterp(showInterp)
            fig.canvas.draw()

    #-----------------------------------------------
    elif artistLabel == 'curve':
        if key_shift:
            coordPoint = [event.mouseevent.xdata, event.mouseevent.ydata]
            if event.artist == curve1:
                if vline1 != None:
                    vline1.set_data([coordPoint[0], coordPoint[0]], [0,1])
                else:
                    vline1 = axs[0].axvline(coordPoint[0], color=pointerColor, alpha=0.5, linestyle='--', linewidth=1, label='vline')
            elif event.artist == curve2:
                if vline2 != None:
                    vline2.set_data([coordPoint[0], coordPoint[0]], [0,1])
                else:
                    vline2 = axs[1].axvline(coordPoint[0], color=pointerColor, alpha=0.5, linestyle='--', linewidth=1, label='vline')
            fig.canvas.draw()

    #-----------------------------------------------
    elif artistLabel == 'points':
        if key_control:
            ind = event.ind[0]
            if event.artist == points1:
                coordPoint = [x1[ind], y1[ind]]
                if vline1 != None:
                    vline1.set_data([coordPoint[0], coordPoint[0]], [0,1])
                else:
                    vline1 = axs[0].axvline(coordPoint[0], color=pointerColor, linestyle='--', linewidth=1, label='vline')
            elif event.artist == points2:
                coordPoint = [x2[ind], y2[ind]]
                if vline2 != None:
                    vline2.set_data([coordPoint[0], coordPoint[0]], [0,1])
                else:
                    vline2 = axs[1].axvline(coordPoint[0], color=pointerColor, linestyle='--', linewidth=1, label='vline')
            fig.canvas.draw()

#=========================================================================================
def zoom(event):
    
    if event.inaxes in axs:
        scale_zoom = 1.2
        cur_xlim = event.inaxes.get_xlim()
        cur_ylim = event.inaxes.get_ylim()
        xdata = event.xdata # get event x location
        ydata = event.ydata # get event y location
        if event.button == 'up':                        # zoom in
            scale_factor = 1 / scale_zoom
        elif event.button == 'down':                    # zoom out
            scale_factor = scale_zoom
        else:
            # deal with something that should never happen
            scale_factor = 1
        new_width = (cur_xlim[1] - cur_xlim[0]) * scale_factor
        new_height = (cur_ylim[1] - cur_ylim[0]) * scale_factor
        relx = (cur_xlim[1] - xdata)/(cur_xlim[1] - cur_xlim[0])
        rely = (cur_ylim[1] - ydata)/(cur_ylim[1] - cur_ylim[0])
        
        event.inaxes.set_xlim(xdata - new_width * (1-relx), xdata + new_width * (relx))
        event.inaxes.set_ylim(ydata - new_height * (1-rely), ydata + new_height * (rely))
        
        updateConnections()
        event.inaxes.figure.canvas.draw()
    
#=========================================================================================
def updateAxes():
    # Full vertical range on both plots, horizontal range set from pointers 
    
    # autoscale to get ylim range
    axs[0].relim()
    axs[1].relim()
    axs[0].autoscale(axis='y')
    axs[1].autoscale(axis='y')
    ylim_axs0 = axs[0].get_ylim()
    ylim_axs1 = axs[1].get_ylim()
    
    # hide 
    displayInterp(False)
    linecursor1.set_visible(False)
    linecursor2.set_visible(False)

    # autoscale to get xlim range only from pointers
    curve1.set_visible(False)
    curve2.set_visible(False)
    axs[0].relim(visible_only=True)
    axs[1].relim(visible_only=True)
    axsInterp.relim(visible_only=True)
    axs[0].autoscale()
    axs[1].autoscale()
    axsInterp.autoscale()
    xlim_axs0 = axs[0].get_xlim()
    xlim_axs1 = axs[1].get_xlim()
    
    # apply xlim and ylim ranges
    curve1.set_visible(True)
    curve2.set_visible(True)
    axs[0].set_xlim(xlim_axs0)
    axs[1].set_xlim(xlim_axs1)
    axs[0].set_ylim(ylim_axs0)
    axs[1].set_ylim(ylim_axs1)
    
    updateConnections()
    displayInterp(showInterp)
    fig.canvas.draw()

#=========================================================================================
def onKeyPress(event):
    global key_x, key_shift, key_control, vline1, vline2, artistsList_Dict
    global showInterp

    sys.stdout.flush()

    #-----------------------------------------------
    if event.key == 'A':
        # Full vertical range on the current plot

        if event.inaxes not in axs: return

        # hide
        linecursor1.set_visible(False)
        linecursor2.set_visible(False)

        # autoscale only on y
        event.inaxes.relim()
        event.inaxes.autoscale(axis='y')
        updateConnections()
        event.inaxes.figure.canvas.draw()

    #-----------------------------------------------
    if event.key == 'a':
        updateAxes()

    #-----------------------------------------------
    if event.key == 'x':
        key_x = True

    #-----------------------------------------------
    if event.key == 'X':
        deleteConnections()
        deleteInterp()
        showInterp = False
        displayInterp(showInterp)
        fig.canvas.draw()

    #-----------------------------------------------
    if event.key == 'c':
        if vline1 != None and vline2 != None :
            coordX1 = float(vline1.get_xdata()[0])
            coordX2 = float(vline2.get_xdata()[0])
            # current coordsX1, coordsX2. Will be defined later from setInterp
            coordsX1_cur = sorted([float(line.get_xdata()[0]) for line in vline1List])
            coordsX2_cur = sorted([float(line.get_xdata()[0]) for line in vline2List])
            # Check positions
            if np.searchsorted(coordsX1_cur, coordX1) != np.searchsorted(coordsX2_cur, coordX2):
                print("Error: Connection not possible because it would cross existing connections") 
                return
            
            connect = ConnectionPatch(color=pointerColor, alpha=0.5, linewidth=1, picker=5, clip_on=True, label='connection',
                        xyA=(coordX1, axs[0].get_ylim()[0]), coordsA=axs[0].transData,
                        xyB=(coordX2, axs[1].get_ylim()[1]), coordsB=axs[1].transData)
            fig.add_artist(connect)
            artistsList_Dict[id(connect)] = [connect, vline1, vline2]
            vline1List.append(vline1)
            vline2List.append(vline2)
            vline1 = None
            vline2 = None
        
            if len(vline1List) >= 2: 
                setInterp()
                displayInterp(showInterp)
            fig.canvas.draw()

    #-----------------------------------------------
    elif event.key == 'h':
        print(interactions)

    #-----------------------------------------------
    elif event.key == 'p':

        counterFilename = 1
        fileNameTemplate = 'savePyAnalySeries_pdfFile_{}.pdf'
        while os.path.isfile(fileNameTemplate.format("%02d" %counterFilename)):
            counterFilename += 1
        fileName = fileNameTemplate.format("%02d" %counterFilename)
        plt.savefig(fileName)
        print("Info: saved pdf in file ", fileName)

        counterFilename = 1
        fileNameTemplate = 'savePyAnalySeries_pngFile_{}.png'
        while os.path.isfile(fileNameTemplate.format("%02d" %counterFilename)):
            counterFilename += 1
        fileName = fileNameTemplate.format("%02d" %counterFilename)
        plt.savefig(fileName)
        print("Info: saved png in file ", fileName)

    #-----------------------------------------------
    elif event.key == 's':

        counterFilename = 1
        fileNameTemplate = 'savePyAnalySeries_dataFile_{}.xlsx'
        while os.path.isfile(fileNameTemplate.format("%02d" %counterFilename)):
            counterFilename += 1
        fileName = fileNameTemplate.format("%02d" %counterFilename)

        with pd.ExcelWriter(fileName) as writer:
            df = pd.DataFrame({x1Name: x1, y1Name: y1, x2Name: x2, y2Name: y2, 
                                y2Name + ' interpolated (' + kindInterpolation + ') on ' + x1Name: x2Interp})
            df.to_excel(writer, sheet_name='Data', index=False, float_format="%.8f")
            worksheet = writer.sheets['Data']
            for i, col in enumerate(df.columns, 1): 
                worksheet.column_dimensions[get_column_letter(i)].width = 25

            df = pd.DataFrame({'Coordinates sX1': coordsX1, 'Coordinates X2': coordsX2})
            df.to_excel(writer, sheet_name='Pointers', index=False, float_format="%.8f")
            worksheet = writer.sheets['Pointers']
            for i, col in enumerate(df.columns, 1): 
                worksheet.column_dimensions[get_column_letter(i)].width = 25

        print("Info: saved data in file ", fileName)

    #-----------------------------------------------
    elif event.key == 'z':
        showInterp = not showInterp
        setInterp()
        displayInterp(showInterp)

    #-----------------------------------------------
    elif event.key == 'shift':
        key_shift = True

    #-----------------------------------------------
    elif event.key == 'control':
        key_control = True
        if event.inaxes == axs[0]:
            points1.set_visible(True)
        elif event.inaxes == axs[1]:
            points2.set_visible(True)
        fig.canvas.draw()

#=========================================================================================
def onKeyRelease(event):
    global key_x, key_shift, key_control

    sys.stdout.flush()

    #-----------------------------------------------
    if event.key == 'x':
        key_x = False

    #-----------------------------------------------
    elif event.key == 'shift':
        key_shift = False 

    #-----------------------------------------------
    elif event.key == 'control':
        key_control = False 
        if event.inaxes == axs[0]:
            points1.set_visible(False)
        elif event.inaxes == axs[1]:
            points2.set_visible(False)
        fig.canvas.draw()

#=========================================================================================
def onPress(event):
    global cur_xlim, cur_ylim, press, xpress, ypress, mousepress, press_origin

    if event.inaxes not in axs: return

    #-----------------------------------------------
    if event.button == 3:
        mousepress = 'right'
        press_origin = event.inaxes

    #-----------------------------------------------
    elif event.button == 1:
        mousepress = 'left'
        press_origin = event.inaxes

    cur_xlim = event.inaxes.get_xlim()
    cur_ylim = event.inaxes.get_ylim()
    press = event.xdata, event.ydata
    xpress, ypress = press

#=========================================================================================
def onRelease(event):
    global press
    press = None

#=========================================================================================
def onMotion(event):
    global cur_xlim, cur_ylim, press

    #-----------------------------------------------
    if event.inaxes not in axs:
        press = None
        linecursor1.set_visible(False)
        linecursor2.set_visible(False)
        fig.canvas.draw()
        return

    #-----------------------------------------------
    if event.inaxes is axs[0]:
        linecursor1.set_visible(True)
        linecursor2.set_visible(False)
        linecursor1.set_xdata([event.xdata])
    elif event.inaxes is axs[1]:
        linecursor1.set_visible(False)
        linecursor2.set_visible(True)
        linecursor2.set_xdata([event.xdata])
    event.inaxes.figure.canvas.draw()

    #-----------------------------------------------
    if press is None: return

    #-----------------------------------------------
    # When mousepress has been done not in the listen axe
    if event.inaxes.get_label() != press_origin.get_label(): return

    #-----------------------------------------------
    if mousepress == 'left':
        linecursor1.set_visible(False)
        linecursor2.set_visible(False)
        dx = event.xdata - xpress
        dy = event.ydata - ypress
        cur_xlim -= dx
        cur_ylim -= dy
        event.inaxes.set_xlim(cur_xlim[0], cur_xlim[1])
        event.inaxes.set_ylim(cur_ylim[0], cur_ylim[1])

    #-----------------------------------------------
    elif mousepress == 'right':
        linecursor1.set_visible(False)
        linecursor2.set_visible(False)
        dx = event.xdata - xpress
        dy = event.ydata - ypress
        zoomFactor = 1.2 
        event.inaxes.set_xlim(cur_xlim[0] + dx * zoomFactor, cur_xlim[1] - dx * zoomFactor)
        event.inaxes.set_ylim(cur_ylim[0] + dy * zoomFactor, cur_ylim[1] - dy * zoomFactor)
       
        if event.inaxes == axs[1]:
            axsInterp.set_ylim(cur_ylim[0] + dy * zoomFactor, cur_ylim[1] - dy * zoomFactor)

    updateConnections()
    event.inaxes.figure.canvas.draw()
    event.inaxes.figure.canvas.flush_events()
    
##########################################################################################

#=========================================================================================
loadData(fileData)

#=========================================================================================
fig, axs = plt.subplots(2, 1, figsize=(10,8), num='PyAnalySeries ' + version)

#=========================================================================================
curve1, = axs[0].plot(x1, y1, color=curve1Color, picker=True, pickradius=20, linewidth=curveWidth, label='curve') 
points1 = axs[0].scatter(x1, y1, s=5, marker='o', color=curve1Color, picker=True, pickradius=20, label='points')
points1.set_visible(False)
linecursor1 = axs[0].axvline(color='k', alpha=0.25, linewidth=1)
axs[0].grid(visible=True, which='major', color='lightgray', linestyle='dashed', linewidth=0.5)
axs[0].set_xlabel(x1Name, color=curve1Color)
axs[0].set_ylabel(y1Name, color=curve1Color)
axs[0].patch.set_alpha(0)
axs[0].autoscale()
axs[0].set_label('curve1')

#=========================================================================================
curve2, = axs[1].plot(x2, y2, color=curve2Color, picker=True, pickradius=20, linewidth=curveWidth, label='curve')
points2 = axs[1].scatter(x2, y2, s=5, marker='o', color=curve2Color, picker=True, pickradius=20, label='points')
points2.set_visible(False)
linecursor2 = axs[1].axvline(color='k', alpha=0.25, linewidth=1)
axs[1].grid(visible=True, which='major', color='lightgray', linestyle='dashed', linewidth=0.5)
axs[1].set_xlabel(x2Name, color=curve2Color)
axs[1].set_ylabel(y2Name, color=curve2Color)
axs[1].autoscale()
axs[1].set_label('curve2')

#=========================================================================================
axsInterp = axs[0].twinx()
axsInterp.sharey(axs[1])
axsInterp.set_ylabel(y2Name, color=curve2Color)
axsInterp.set_zorder(-10)
axsInterp.set_visible(showInterp)
axsInterp.set_label('curve2Interp')

#=========================================================================================
fig.canvas.mpl_connect('key_press_event', onKeyPress)
fig.canvas.mpl_connect('key_release_event', onKeyRelease)
fig.canvas.mpl_connect('button_press_event',onPress)
fig.canvas.mpl_connect('button_release_event',onRelease)
fig.canvas.mpl_connect('motion_notify_event',onMotion)
fig.canvas.mpl_connect('pick_event', onPick)
fig.canvas.mpl_connect('scroll_event', zoom)

#=========================================================================================
drawConnections()
setInterp()
updateAxes()

#=========================================================================================
plt.show()
