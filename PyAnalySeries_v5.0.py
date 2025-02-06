#!/usr/bin/env python

#========================================================================================
# Author: Patrick Brockmann CEA/DRF/LSCE - November 2024
#========================================================================================

import sys
import os

import numpy as np
import pandas as pd

from PyQt5.QtWidgets import * 
from PyQt5.QtCore import * 
from PyQt5.QtGui import *

from resources.misc import *
from resources.CustomQColorDialog import CustomQColorDialog 

from resources.displaySingleSerieWindow import displaySingleSerieWindow
from resources.displayOverlaidSeriesWindow import displayOverlaidSeriesWindow
from resources.displayStackedSeriesWindow import displayStackedSeriesWindow

from resources.defineFilterWindow import defineFilterWindow
from resources.displayFilterWindow import displayFilterWindow

from resources.defineInterpolationWindow import defineInterpolationWindow
from resources.displayInterpolationWindow import displayInterpolationWindow

from resources.importDataWindow import importDataWindow

from resources.defineInsolationWindow import defineInsolationWindow

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

#========================================================================================
if len(sys.argv[1:]) >= 1:
    filesWS = sys.argv[1:]
else:
    filesWS = None

#========================================================================================
version = 'v5.0'

open_ws = {}
open_displayWindows = {} 
open_filterWindows = {} 
open_interpolationWindows = {} 
open_importWindow = {}
open_insolationWindow= {}

#========================================================================================
def colorize_item(item, color_name):
    tree_widget.blockSignals(True)
    color = QColor(color_name)
    brush = QBrush(color)
    for col in range(tree_widget.columnCount()):
        item.setBackground(col, brush)
    tree_widget.blockSignals(False)

#========================================================================================
def populate_tree_widget(fileName, itemDict_list):
    global open_ws

    tree_widget.blockSignals(True)

    ws_icon = QIcon("resources/icon_folder.png")

    ws_item = QTreeWidgetItem(tree_widget)
    ws_item.setIcon(0, ws_icon)
    ws_item.setText(0, fileName)
    ws_item.setToolTip(0, fileName)
    ws_item.setExpanded(True)
    ws_item.setFlags(ws_item.flags() & ~Qt.ItemIsSelectable)
    open_ws[id(ws_item)] = ws_item.text(0)

    for itemDict in itemDict_list:
        add_item_tree_widget(ws_item, itemDict)

    unmark_ws(ws_item)

    tree_widget.blockSignals(False)

    return ws_item

#========================================================================================
def add_item_tree_widget(ws_item, itemDict, position=None):

    icon_serie = QIcon("resources/icon_document.png")
    icon_serieDuplicated = QIcon("resources/icon_copy.png")
    icon_filter = QIcon("resources/icon_filter.png")
    icon_interpolate = QIcon("resources/icon_interpolate.png")

    item = QTreeWidgetItem()
    item.setFlags(item.flags() & ~Qt.ItemIsDropEnabled)

    if itemDict['Type'].startswith('Serie'):
        Serie = itemDict['Serie']
        if Serie.index.duplicated().any():
            item.setIcon(0, icon_serieDuplicated)
        else:
            item.setIcon(0, icon_serie)
    elif itemDict['Type'] == 'FILTER':
            item.setIcon(0, icon_filter)
    elif itemDict['Type'] == 'INTERPOLATION':
            item.setIcon(0, icon_interpolate)
    else:
        #print("Error: Type unknown")
        return

    # add item at current_index
    if not ws_item:
        current_index = tree_widget.currentItem()
        if not current_index.parent():
            ws_item = current_index
        else:
            ws_item = current_index.parent()

    if position != None:
        ws_item.insertChild(position, item)
    else:
        ws_item.addChild(item)

    mark_ws(ws_item)            # Mark as to be saved

    item.setData(0, Qt.UserRole, itemDict)

    item.setText(0, itemDict['Name'])
    item.setText(1, itemDict['Id'])
    item.setText(2, itemDict['Type'])

    fontMono = QFont('Monospace', 12)
    item.setFont(1, fontMono)             # format Id

    if itemDict['Type'] == 'INTERPOLATION':
        item.setText(3, itemDict['X1Name'])
        item.setFont(3, fontMono)

    if not itemDict['Type'].startswith('Serie'):
        return

    item.setText(3, itemDict['X'])
    item.setText(4, itemDict['Y'])
    item.setFont(3, fontMono)
    item.setFont(4, fontMono)

    buttonColor = QPushButton()
    buttonColor.setFixedSize(40, 15)
    buttonColor.setStyleSheet(f"background-color: {itemDict['Color']}; border: none; border-radius: 3px;")
    buttonColor.clicked.connect(lambda: selectColor(buttonColor, item))
    tree_widget.setItemWidget(item, 5, buttonColor)

    checkboxInverted = QCheckBox()
    checkboxInverted.setFixedSize(60, 15)
    checkboxInverted.setChecked(itemDict["Y axis inverted"])
    checkboxInverted.stateChanged.connect(lambda: checkboxInverted_changed(checkboxInverted, item))
    tree_widget.setItemWidget(item, 6, checkboxInverted)

#========================================================================================
def on_item_changed(item, column):
    global open_ws

    if not item.parent():
        if column !=0: return
        new_wsName = item.text(0).replace(' *', '')
        old_wsName = open_ws[id(item)]
        if new_wsName == old_wsName: 
            remark_ws(item)
            return
        if new_wsName in open_ws.values():
            QMessageBox.warning(main_window, "Duplicate WS name", f"The ws '{new_wsName}' is already in use. Please choose a unique name.")
            item.setText(0, old_wsName)
        else:
            open_ws[id(item)] = new_wsName
            mark_ws(item)
    else: 
        itemDict = item.data(0, Qt.UserRole)
        itemDict['Name'] = item.text(0)
        if 'X' in itemDict.keys(): itemDict['X'] = item.text(3)
        if 'X1Name' in itemDict.keys(): itemDict['X1Name'] = item.text(3)
        if 'Y' in itemDict.keys(): itemDict['Y'] = item.text(4)
        item.setData(0, Qt.UserRole, itemDict)
        update_items_from_data(item)

#========================================================================================
def selectColor(buttonColor, serie_item):
    serieDict = serie_item.data(0, Qt.UserRole)
    starting_color = serieDict['Color']
    color = CustomQColorDialog.getColor(starting_color)
    if color:
        serieDict = serieDict | {'Color': color.name()}
        serie_item.setData(0, Qt.UserRole, serieDict)
        update_items_from_data(serie_item)

#========================================================================================
def checkboxInverted_changed(checkboxInverted, serie_item):
    serieDict = serie_item.data(0, Qt.UserRole)
    serieDict = serieDict | {'Y axis inverted': checkboxInverted.isChecked()}
    serie_item.setData(0, Qt.UserRole, serieDict)
    update_items_from_data(serie_item)

#========================================================================================
def update_items_from_data(ref_item):

    ref_itemDict = ref_item.data(0, Qt.UserRole)

    allItems = tree_widget.get_children()

    for item in allItems:
        itemDict = item.data(0, Qt.UserRole)
        if itemDict['Id'] == ref_itemDict['Id']:
            if  itemDict['Type'].startswith('Serie'):
                buttonColor = tree_widget.itemWidget(item, 5)
                if buttonColor: buttonColor.setStyleSheet(f"background-color: {ref_itemDict['Color']}; border: none; border-radius: 3px;")
                checkboxInverted = tree_widget.itemWidget(item, 6)
                if checkboxInverted: checkboxInverted.setChecked(ref_itemDict["Y axis inverted"])
                sync_window_with_item(item)
            item.setText(0, ref_itemDict['Name'])
            if 'X' in ref_itemDict.keys(): item.setText(3, ref_itemDict['X'])
            if 'X1Name' in ref_itemDict.keys(): item.setText(3, ref_itemDict['X1Name'])
            if 'Y' in ref_itemDict.keys(): item.setText(4, ref_itemDict['Y'])
            item.setData(0, Qt.UserRole, ref_itemDict)
            if ref_item.parent() == item.parent():
                mark_ws(item.parent())

#========================================================================================
def sync_window_with_item(item):
    itemDict = item.data(0, Qt.UserRole)
    Id_window = itemDict['Id']
    for key in open_displayWindows.keys():
        displayWindow = open_displayWindows[key]
        displayWindow.sync_with_item(item)
    for key in open_filterWindows.keys():
        filterWindow = open_filterWindows[key]
        filterWindow.sync_with_item(item)
    for key in open_interpolationWindows.keys():
        interpolationWindow = open_interpolationWindows[key]
        interpolationWindow.sync_with_item(item)

#========================================================================================
def load_WorkSheet(fileName):

    if not os.path.exists(fileName): return

    if fileName in open_ws.values():
        msg = f'{fileName} already loaded'
        main_window.statusBar().showMessage(msg, 5000)
        return 

    main_window.statusBar().showMessage(fileName + ' loading', 5000)
    QApplication.processEvents()

    #--------------------------------------------------------------------
    sheetNames = pd.read_excel(fileName, sheet_name=None).keys()

    itemDict_list = []

    for sheetName in sheetNames:

        #-------------------------------------
        if sheetName.startswith('Serie Id-'):

            try:
                df = pd.read_excel(fileName, sheet_name=sheetName, na_filter=False)

                color = QColor(df['Color'][0])
                if color.isValid():
                    Color = df['Color'][0]
                else:
                    Color = generate_color()

                serieDict = {
                    'Id': 'Id-' + sheetName.split('Serie Id-')[1],
                    'Type': df['Type'][0],
                    'Name': df['Name'][0],
                    'X':  df.columns[0],
                    'Y':  df.columns[1],
                    'Y axis inverted': bool(df['Y axis inverted'][0]),
                    'Color': Color,
                    'Comment': df['Comment'][0],
                    'History': df['History'][0],
                    'Serie': pd.Series(cleanList(df.iloc[:,1]), index=cleanList(df.iloc[:,0]))
                }

                if 'InterpolationMode' in df.columns:
                    serieDict = serieDict | {
                        'InterpolationMode': df['InterpolationMode'][0],
                        'X1Coords': cleanList(df['X1Coords']),
                        'X2Coords': cleanList(df['X2Coords']),
                        'XOriginal': df.columns[11],
                        'XOriginalValues': cleanList(df.iloc[:,11])
                    }

                itemDict_list.append(serieDict)

            except:
                serieDict = None
                msg = f"The file '{fileName}' contains a serie that is wrongly formatted in {sheetName} sheet."
                main_window.statusBar().showMessage(msg, 5000)
                QApplication.processEvents()

        #-------------------------------------
        elif sheetName.startswith('FILTER Id-'):
            
            try:
                df = pd.read_excel(fileName, sheet_name=sheetName, na_filter=False)
                filterDict = {
                        'Id': 'Id-' + sheetName.split('FILTER Id-')[1],
                        'Type': df['Type'][0],
                        'Name': df['Name'][0],
                        'Parameters': str(df['Parameters'][0]),
                        'Comment': df['Comment'][0],
                        'History': df['History'][0]
                }

                itemDict_list.append(filterDict)

            except:
                filterDict = None
                msg = f"The file '{fileName}' contains a FILTER that is wrongly formatted in {sheetName} sheet."
                main_window.statusBar().showMessage(msg, 5000)
                QApplication.processEvents()

        #-------------------------------------
        elif sheetName.startswith('INTERPOLATION Id-'):

            try:
                df = pd.read_excel(fileName, sheet_name=sheetName, na_filter=False)
                interpolationDict = {
                        'Id': 'Id-' + sheetName.split('INTERPOLATION Id-')[1],
                        'Type': df['Type'][0],
                        'Name': df['Name'][0],
                        'X1Coords': df['X1Coords'].values,
                        'X2Coords': df['X2Coords'].values,
                        'X1Name': df['X1Name'][0],
                        'Comment': df['Comment'][0],
                        'History': df['History'][0]
                }

                itemDict_list.append(interpolationDict)

            except:
                interpolationDict = None
                msg = f"The file '{fileName}' contains an INTERPOLATION that is wrongly formatted in {sheetName} sheet."
                main_window.statusBar().showMessage(msg, 5000)
                QApplication.processEvents()

    #--------------------------------------------------------------------
    populate_tree_widget(fileName, itemDict_list)

    #--------------------------------------------------------------------
    main_window.statusBar().showMessage(fileName + ' loaded', 5000)

#========================================================================================
def new_WorkSheet():

    fileNameTemplate = 'new_{}.xlsx'
    counterFilename = 1
    while fileNameTemplate.format("%02d" %counterFilename) in open_ws.values():
        counterFilename += 1
    fileName = fileNameTemplate.format("%02d" %counterFilename)
    
    ws_item = populate_tree_widget(fileName, [])
    mark_ws(ws_item)
    tree_widget.setCurrentItem(ws_item)
    tree_widget.clearSelection()

#========================================================================================
def open_WorkSheet():

    fileName, _ = QFileDialog.getOpenFileName(main_window, "Open Excel File", "", "Excel Files (*.xlsx)")
    if fileName:
        base_dir = os.getcwd()
        fileName = os.path.relpath(fileName, base_dir)          # get relative path
        series = load_WorkSheet(fileName)

#========================================================================================
def autofit_columns(worksheet):
    for col in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        worksheet.column_dimensions[column_letter].width = max_length + 5 

#========================================================================================
def save_WorkSheet(ws_item):

    outFile = ws_item.text(0).replace(" *", "")

    #-----------------------
    try:
        wb = Workbook()

        for n in range(ws_item.childCount()):

            item = ws_item.child(n)
            itemDict = item.data(0, Qt.UserRole)

            #-----------------------
            if itemDict["Type"].startswith('Serie'):
                sheetName = f'{itemDict["Type"].split(" ")[0]} {itemDict["Id"]}'
                ws = wb.create_sheet(title=sheetName)

                ws.cell(row=1, column=1, value=itemDict['X'])
                ws.cell(row=1, column=2, value=itemDict['Y'])
                ws.cell(row=1, column=3, value='Type')
                ws.cell(row=1, column=4, value='Name')
                ws.cell(row=1, column=5, value='Y axis inverted')
                ws.cell(row=1, column=6, value='Color')
                ws.cell(row=1, column=7, value='Comment')
                ws.cell(row=1, column=8, value='History')

                for i, (index, value) in enumerate(itemDict['Serie'].items(), start=2):
                    ws.cell(row=i, column=1, value=index)
                    ws.cell(row=i, column=2, value=value)
                ws.cell(row=2, column=3, value=itemDict['Type'])
                ws.cell(row=2, column=4, value=itemDict['Name'])
                ws.cell(row=2, column=5, value=itemDict['Y axis inverted'])
                ws.cell(row=2, column=6, value=itemDict['Color'])
                ws.cell(row=2, column=7, value=itemDict['Comment'])
                ws.cell(row=2, column=8, value=itemDict['History'])
    
                if 'InterpolationMode' in itemDict:
                    ws.cell(row=1, column=9, value='InterpolationMode')
                    ws.cell(row=1, column=10, value='X1Coords')
                    ws.cell(row=1, column=11, value='X2Coords')
                    ws.cell(row=1, column=12, value=itemDict['XOriginal'])

                    ws.cell(row=2, column=9, value=itemDict['InterpolationMode'])
                    for i, value in enumerate(itemDict['X1Coords'], start=2):
                        ws.cell(row=i, column=10, value=value)
                    for i, value in enumerate(itemDict['X2Coords'], start=2):
                        ws.cell(row=i, column=11, value=value)
                    for i, value in enumerate(itemDict['XOriginalValues'], start=2):
                        ws.cell(row=i, column=12, value=value)

            #-----------------------
            elif itemDict["Type"] == 'FILTER':
                sheetName = f'{itemDict["Type"]} {itemDict["Id"]}'
                ws = wb.create_sheet(title=sheetName)

                ws.cell(row=1, column=1, value='Type')
                ws.cell(row=1, column=2, value='Name')
                ws.cell(row=1, column=3, value='Parameters')
                ws.cell(row=1, column=4, value='Comment')
                ws.cell(row=1, column=5, value='History')
                
                ws.cell(row=2, column=1, value=itemDict['Type'])
                ws.cell(row=2, column=2, value=itemDict['Name'])
                ws.cell(row=2, column=3, value=itemDict['Parameters'])
                ws.cell(row=2, column=4, value=itemDict['Comment'])
                ws.cell(row=2, column=5, value=itemDict['History'])
                
            #-----------------------
            elif itemDict["Type"] == 'INTERPOLATION':
                sheetName = f'{itemDict["Type"]} {itemDict["Id"]}'
                ws = wb.create_sheet(title=sheetName)

                ws.cell(row=1, column=1, value='X1Coords')
                ws.cell(row=1, column=2, value='X2Coords')
                ws.cell(row=1, column=3, value='X1Name')
                ws.cell(row=1, column=4, value='Type')
                ws.cell(row=1, column=5, value='Name')
                ws.cell(row=1, column=6, value='Comment')
                ws.cell(row=1, column=7, value='History')

                for i, value in enumerate(itemDict['X1Coords'], start=2):
                    ws.cell(row=i, column=1, value=value)
                for i, value in enumerate(itemDict['X2Coords'], start=2):
                    ws.cell(row=i, column=2, value=value)
                ws.cell(row=2, column=3, value=itemDict['X1Name'])
                ws.cell(row=2, column=4, value=itemDict['Type'])
                ws.cell(row=2, column=5, value=itemDict['Name'])
                ws.cell(row=2, column=6, value=itemDict['Comment'])
                ws.cell(row=2, column=7, value=itemDict['History'])
                
        if ws_item.childCount() > 0:
            if 'Sheet' in wb.sheetnames:
                del wb['Sheet']
            for sheet in wb.worksheets:
                autofit_columns(sheet)

        wb.save(outFile)
        return True 

    #-----------------------
    except:
        return False 

#========================================================================================
def save_WorkSheets():

    display_error = False
    for ws_item in tree_widget.get_parents():
        if ws_item.data(0, Qt.UserRole): 
            outFile = ws_item.text(0).replace(" *", "")
            if save_WorkSheet(ws_item):
                unmark_ws(ws_item)
            else:
                main_window.statusBar().showMessage(f'Error when saving {outFile}', 5000)
                display_error = True

    if not display_error:
        main_window.statusBar().showMessage('Worksheets saved', 5000)

#========================================================================================
def import_Series():
    global open_importWindow

    current_index = tree_widget.currentItem()
    if not current_index:
        new_WorkSheet()
    
    Id_importWindow = '123456'

    if open_importWindow:
        importWindow = open_importWindow[Id_importWindow]
        importWindow.raise_()
        importWindow.activateWindow()
    else:
        importWindow = importDataWindow(open_importWindow, add_item_tree_widget)
        open_importWindow[Id_importWindow] = importWindow
        importWindow.show()

#========================================================================================
def define_insolationSerie():
    global open_insolationWindow

    current_index = tree_widget.currentItem()
    if not current_index:
        new_WorkSheet()
    
    Id_insolationWindow = '123456'

    if open_insolationWindow:
        insolationWindow = open_insolationWindow[Id_insolationWindow]
        insolationWindow.raise_()
        insolationWindow.activateWindow()
    else:
        insolationWindow = defineInsolationWindow(open_insolationWindow, add_item_tree_widget)
        open_insolationWindow[Id_insolationWindow] = insolationWindow
        insolationWindow.show()

#========================================================================================
def define_astroSerie():
    return

#========================================================================================
def create_tree_widget():

    fontMono = QFont('Monospace', 12)

    tree_widget = CustomTreeWidget()
    tree_widget.setColumnCount(7)
    tree_widget.setHeaderLabels(["Name", "Id", "Type", "X", "Y", "Color", "Y axis inverted"])
    tree_widget.setColumnWidth(0, 300)
    tree_widget.setColumnWidth(1, 150)
    tree_widget.setColumnWidth(2, 150)
    tree_widget.setColumnWidth(3, 300)
    tree_widget.setColumnWidth(4, 300)
    tree_widget.setColumnWidth(5, 50)
    tree_widget.setColumnWidth(6, 50)
    tree_widget.setTextElideMode(Qt.ElideLeft)
    tree_widget.setSelectionMode(QTreeWidget.ExtendedSelection)
    tree_widget.setIconSize(QSize(16, 16))
    tree_widget.setStyleSheet("""
        QTreeView:selected {
            background-color: lightsteelblue;
            color: black;
        }
    """)
    tree_widget.setFocusPolicy(Qt.ClickFocus)
    tree_widget.headerItem().setFont(1, fontMono)
    tree_widget.headerItem().setFont(3, fontMono)
    tree_widget.headerItem().setFont(4, fontMono)

    delegate = QStyledItemDelegate()
    delegate.createEditor = lambda parent, option, index: (
        QLineEdit(parent, font=fontMono)
    )
    tree_widget.setItemDelegateForColumn(1, delegate)
    tree_widget.setItemDelegateForColumn(3, delegate)
    tree_widget.setItemDelegateForColumn(4, delegate)

    tree_widget.setDragEnabled(True)
    tree_widget.setAcceptDrops(True)
    tree_widget.setDropIndicatorShown(True)
    tree_widget.setDragDropMode(QTreeWidget.InternalMove)
    tree_widget.invisibleRootItem().setFlags(Qt.ItemIsEnabled)      # root non droppable

    return tree_widget


#========================================================================================
class CustomTreeWidget(QTreeWidget):

    def __init__(self):
        super().__init__()
        self.clipboard_items = []

    def dropEvent(self, event):
        dragged_item = self.currentItem()
        target_item = self.itemAt(event.pos())

        if not target_item or dragged_item.parent() != target_item.parent():
            event.ignore()
            return
        
        # Find the position where to move the drag item
        position = target_item.parent().indexOfChild(target_item)
        # Retrieve the data from the dragged item
        itemDict = dragged_item.data(0, Qt.UserRole)
        dragged_item.parent().removeChild(dragged_item)
        # Use the `add_item_tree_widget` function to add the dragged item to the target parent
        add_item_tree_widget(target_item.parent(), itemDict, position)

        # Call the default implementation if the drop is valid
        super().dropEvent(event)

    def get_parents(self):
        parents = []
        for i in range(self.topLevelItemCount()):
            parents.append(self.topLevelItem(i))
        return parents

    def get_children(self):
        children = []
        for i in range(self.topLevelItemCount()):
            parent_item = self.topLevelItem(i)
            for j in range(parent_item.childCount()):
                children.append(parent_item.child(j))
        return children

#========================================================================================
def get_unique_selected_items(tree_widget):
    selected_items = tree_widget.selectedItems()
    unique_ids = set()
    unique_items = []

    for item in selected_items:
        itemDict = item.data(0, Qt.UserRole)

        if itemDict['Id'] not in unique_ids:
            unique_ids.add(itemDict['Id'])  # Ajoute l'ID à l'ensemble
            unique_items.append(item)  # Ajoute l'item à la liste unique

    return unique_items

#========================================================================================
def displaySingleSerie_selected_series():
    global open_displayWindows

    items = get_unique_selected_items(tree_widget)
    if len(items) == 0:
        main_window.statusBar().showMessage('Please select at least 1 serie', 5000)
        return

    for item in items:
        itemDict = item.data(0, Qt.UserRole)

        Id_displayWindow = itemDict['Id']

        if Id_displayWindow in open_displayWindows:
            displayWindow = open_displayWindows[Id_displayWindow]
            displayWindow.raise_()
            displayWindow.activateWindow()
        else:
            if itemDict['Type'].startswith('Serie'): 
                displayWindow = displaySingleSerieWindow(Id_displayWindow, open_displayWindows, item)
            elif itemDict['Type'] == 'FILTER':
                displayWindow = displayFilterWindow(Id_displayWindow, open_displayWindows, item)
            elif itemDict['Type'] == 'INTERPOLATION':
                displayWindow = displayInterpolationWindow(Id_displayWindow, open_displayWindows, item)
            open_displayWindows[Id_displayWindow] = displayWindow
            displayWindow.show()

    #-------------------------------------------------------------
    main_window.setFocus()                  # replace selection
    tree_widget.clearSelection()
    for item in items:
        item.setSelected(True)

#========================================================================================
def displayMultipleSeries_selected_series(overlaid=True):
    global open_displayWindows

    items = get_unique_selected_items(tree_widget)
    if len(items) == 0:
        main_window.statusBar().showMessage('Please select at least 1 serie', 5000)
        return
    elif len(items) == 1:                             # If only 1 item selected
        displaySingleSerie_selected_series()
        return

    items_selected = []                             # select only series
    serieDicts = []
    for item in items:
        itemDict = item.data(0, Qt.UserRole)
        if  itemDict['Type'].startswith('Serie'): 
            items_selected.append(item)
            serieDicts.append(itemDict)

    #-------------------------------------------------------------
    Id_displayWindow = tuple(serieDict['Id'] for serieDict in serieDicts)

    if Id_displayWindow in open_displayWindows:
        displayWindow = open_displayWindows[Id_displayWindow]
        displayWindow.raise_()
        displayWindow.activateWindow()
    else:
        if overlaid:
            displayWindow = displayOverlaidSeriesWindow(Id_displayWindow, open_displayWindows, items_selected)
        else:
            displayWindow = displayStackedSeriesWindow(Id_displayWindow, open_displayWindows, items_selected)
        open_displayWindows[Id_displayWindow] = displayWindow
        displayWindow.show()

    #-------------------------------------------------------------
    main_window.setFocus()                  # replace selection
    tree_widget.clearSelection()
    for item in items_selected:
        item.setSelected(True)

#========================================================================================
def define_filter():
    global open_filterWindows

    items = get_unique_selected_items(tree_widget)
    items_selected = []                             # select only series
    for item in items:
        serieDict = item.data(0, Qt.UserRole)
        if  serieDict['Type'].startswith('Serie'): 
            items_selected.append(item)

    if len(items_selected) != 1 : 
        main_window.statusBar().showMessage('Please select only 1 serie', 5000)
        return

    #-------------------------------------------------------------
    Id_filterWindow = generate_Id()

    if Id_filterWindow in open_filterWindows:
        filterWindow = open_filterWindows[Id_filterWindow]
        filterWindow.raise_()
        filterWindow.activateWindow()
    else:
        filterWindow = defineFilterWindow(Id_filterWindow, open_filterWindows, item, add_item_tree_widget)
        open_filterWindows[Id_filterWindow] = filterWindow
        filterWindow.show()

    #-------------------------------------------------------------
    main_window.setFocus()                  # replace selection
    tree_widget.clearSelection()
    item.setSelected(True)

#========================================================================================
def apply_filter():

    items = get_unique_selected_items(tree_widget)
    itemSeries_selected = []
    itemFilters_selected = []
    for item in items:
        itemDict = item.data(0, Qt.UserRole)
        if  itemDict['Type'].startswith('Serie'): 
            itemSeries_selected.append(item)
        elif itemDict['Type'] == 'FILTER':
            itemFilters_selected.append(item)

    if len(itemFilters_selected) != 1 or len(itemSeries_selected) < 1:
        main_window.statusBar().showMessage('Please select 1 FILTER and at least 1 serie', 5000)
        return
       
    #-------------------------------------------------------------
    tree_widget.clearSelection()
    itemFilter = itemFilters_selected[0]
    colorize_item(itemFilter, 'red')
    for item in itemSeries_selected:
        colorize_item(item, 'green')

    reply = QMessageBox.question(
        main_window, 
        "Apply filter confirmation",
        "Do you want to apply filter on selected series ?",
        QMessageBox.Yes | QMessageBox.No,
        QMessageBox.No
    )
    
    if reply == QMessageBox.No:
        for item in items:
            colorize_item(item, 'white')
        tree_widget.clearSelection()
        return

    #-------------------------------------------------------------
    filterDict = itemFilter.data(0, Qt.UserRole)
    filter_window_size = int(filterDict['Parameters'])

    for item in itemSeries_selected:
        serieDict = item.data(0, Qt.UserRole)
        serie = serieDict['Serie']
        serie = serie.groupby(serie.index).mean()

        filtered_Id = generate_Id()
        filtered_serieDict = serieDict | {'Id': filtered_Id,
            'Type': 'Serie filtered',
            'Serie': defineFilterWindow.moving_average(serie, window_size=filter_window_size),
            'Color': generate_color(exclude_color=serieDict['Color']),
            'History': append_to_htmlText(serieDict['History'], 
                f'serie <i><b>{serieDict["Id"]}</i></b> filtered with FILTER <i><b>{filterDict["Id"]}</i></b> with a moving average of size {filter_window_size}<BR>---> serie <i><b>{filtered_Id}</b></i>'),
            'Comment': ''
        }
        ws_item = item.parent()
        position = ws_item.indexOfChild(item)
        add_item_tree_widget(ws_item, filtered_serieDict, position+1)

    #-------------------------------------------------------------
    main_window.setFocus()                  # replace selection
    tree_widget.clearSelection()
    for item in itemSeries_selected + itemFilters_selected:
        colorize_item(item, 'white')
        item.setSelected(True)

#========================================================================================
def define_interpolation():
    global open_interpolationWindows

    items = get_unique_selected_items(tree_widget)
    itemSeries_selected = []
    itemInterpolations_selected = []
    for item in items:
        itemDict = item.data(0, Qt.UserRole)
        if  itemDict['Type'].startswith('Serie'): 
            itemSeries_selected.append(item)
        elif itemDict['Type'] == 'INTERPOLATION':
            itemInterpolations_selected.append(item)

    if not (len(itemInterpolations_selected) <= 1 and len(itemSeries_selected) >= 2):
        main_window.statusBar().showMessage('Please select at least 2 series and optionnaly 1 INTERPOLATION', 5000)
        return
       
    #-------------------------------------------------------------
    if len(itemInterpolations_selected) == 1:
        itemInterpolation = itemInterpolations_selected[0]
        itemDict = itemInterpolation.data(0, Qt.UserRole)
        Id_interpolationWindow = itemDict['Id']
    else:
        itemInterpolation = None
        Id_interpolationWindow = generate_Id()
    items = itemSeries_selected

    if Id_interpolationWindow in open_interpolationWindows:
        interpolationWindow = open_interpolationWindows[Id_interpolationWindow]
        interpolationWindow.raise_()
        interpolationWindow.activateWindow()
    else:
        interpolationWindow = defineInterpolationWindow(Id_interpolationWindow, open_interpolationWindows, 
                                itemInterpolation, items, add_item_tree_widget)
        open_interpolationWindows[Id_interpolationWindow] = interpolationWindow
        interpolationWindow.show()

    #-------------------------------------------------------------
    main_window.setFocus()                  # replace selection
    tree_widget.clearSelection()
    for item in itemSeries_selected + itemInterpolations_selected:
        colorize_item(item, 'white')
        item.setSelected(True)

#========================================================================================
def apply_interpolation(interpolationMode):

    items = get_unique_selected_items(tree_widget)
    itemSeries_selected = []
    itemInterpolations_selected = []
    for item in items:
        itemDict = item.data(0, Qt.UserRole)
        if  itemDict['Type'].startswith('Serie'): 
            itemSeries_selected.append(item)
        elif itemDict['Type'] == 'INTERPOLATION':
            itemInterpolations_selected.append(item)

    if len(itemInterpolations_selected) != 1 or len(itemSeries_selected) < 1:
        main_window.statusBar().showMessage('Please select 1 INTERPOLATION and at least 1 serie', 5000)
        return
       
    #-------------------------------------------------------------
    tree_widget.clearSelection()
    itemInterpolation = itemInterpolations_selected[0]
    colorize_item(itemInterpolation, 'red')
    for item in itemSeries_selected:
        colorize_item(item, 'green')

    reply = QMessageBox.question(
        main_window, 
        "Apply interpolation confirmation",
        "Do you want to apply interpolation on selected series ?",
        QMessageBox.Yes | QMessageBox.No,
        QMessageBox.No
    )
    
    if reply == QMessageBox.No:
        for item in items:
            colorize_item(item, 'white')
        tree_widget.clearSelection()
        return

    #-------------------------------------------------------------
    interpolationDict = itemInterpolation.data(0, Qt.UserRole)
    X1Coords = interpolationDict['X1Coords']
    X2Coords = interpolationDict['X2Coords']
    f_1to2, f_2to1 = defineInterpolationWindow.defineInterpolationFunctions(X1Coords, X2Coords, interpolationMode=interpolationMode)

    for item in itemSeries_selected:
        serieDict = item.data(0, Qt.UserRole)
        serie = serieDict['Serie']
        serie = serie.groupby(serie.index).mean()

        interpolated_Id = generate_Id()
        interpolated_serieDict = serieDict | {'Id': interpolated_Id,
            'Type': 'Serie interpolated',
            'Serie': pd.Series(serie.values, index=f_2to1(serie.index)),
            'InterpolationMode': interpolationMode,
            'X': interpolationDict['X1Name'], 
            'XOriginal': serieDict['X'], 
            'XOriginalValues': serie.index.to_list(),
            'X1Coords': X1Coords,
            'X2Coords': X2Coords, 
            'Color': generate_color(exclude_color=serieDict['Color']),
            'History': append_to_htmlText(serieDict['History'], 
                f'serie <i><b>{serieDict["Id"]}</i></b> interpolated with INTERPOLATION <i><b>{interpolationDict["Id"]}</i></b> with mode {interpolationMode}<BR>---> serie <i><b>{interpolated_Id}</b></i>'),
            'Comment': ''
        }

        ws_item = item.parent()
        position = ws_item.indexOfChild(item)
        add_item_tree_widget(ws_item, interpolated_serieDict, position+1)

    #-------------------------------------------------------------
    main_window.setFocus()                  # replace selection
    tree_widget.clearSelection()
    for item in itemSeries_selected + itemInterpolations_selected:
        colorize_item(item, 'white')
        item.setSelected(True)

#========================================================================================
def close_all_windows():
    global open_displayWindows

    for Id_displayWindow in list(open_displayWindows.keys()):
        open_displayWindows[Id_displayWindow].close()
    open_displayWindows.clear()

#========================================================================================
def delete_parent_node(item):
    global open_ws

    index = tree_widget.indexOfTopLevelItem(item)
    tree_widget.takeTopLevelItem(index)
    del open_ws[id(item)]

#========================================================================================
def show_context_menu(point):
    item = tree_widget.itemAt(point)
    if item and not item.parent():  # Only allow delete on parents (ws)
        context_menu = QMenu(tree_widget)
        delete_action = context_menu.addAction("Remove")
        action = context_menu.exec_(tree_widget.mapToGlobal(point))
        if action == delete_action:
            delete_parent_node(item)

#========================================================================================
def is_item_in_ws(ws_item, child_item):
    child_itemDict = child_item.data(0, Qt.UserRole)
    for i in range(ws_item.childCount()):
        item = ws_item.child(i)
        itemDict = item.data(0, Qt.UserRole)
        if child_itemDict['Id'] == itemDict['Id']:
            return True
    return False

#========================================================================================
def mark_ws(ws_item):
    tree_widget.blockSignals(True)
    ws_item.setData(0, Qt.UserRole, True)  # Mark ws_item as modified
    if not ws_item.text(0).endswith(' *'):
        ws_item.setText(0, f"{ws_item.text(0)} *") 
    tree_widget.blockSignals(False)

#========================================================================================
def unmark_ws(ws_item):
    tree_widget.blockSignals(True)
    ws_item.setData(0, Qt.UserRole, False)  # Reset modification state
    ws_item.setText(0, ws_item.text(0).replace(" *", ""))  # Remove visual cue
    tree_widget.blockSignals(False)
    
#========================================================================================
def remark_ws(ws_item):
    tree_widget.blockSignals(True)
    if ws_item.data(0, Qt.UserRole) and not ws_item.text(0).endswith(' *'):
        ws_item.setText(0, f"{ws_item.text(0)} *") 
    tree_widget.blockSignals(False)

#========================================================================================
def copy_items():
    selected_items = tree_widget.selectedItems()
    tree_widget.clipboard_items = selected_items

#========================================================================================
def cut_items():
    selected_items = tree_widget.selectedItems()
    tree_widget.clipboard_items = selected_items

    for item in selected_items:
        ws_item = item.parent()
        ws_item.removeChild(item)
        mark_ws(ws_item)

#========================================================================================
def paste_items():
    target_item = tree_widget.currentItem()
    ws_item = target_item.parent() if target_item.parent() else target_item

    for item in tree_widget.clipboard_items:
        if is_item_in_ws(ws_item, item):
            main_window.statusBar().showMessage('Item already in', 5000)
            continue
        itemDict = item.data(0, Qt.UserRole)
        add_item_tree_widget(ws_item, itemDict)

#========================================================================================
def on_item_double_clicked(item, column):

    tree_widget.blockSignals(True)

    if not item.parent(): item_isWS = True
    else: item_isWS = False

    itemDict = item.data(0, Qt.UserRole)

    if column == 0:
        item.setFlags(item.flags() | Qt.ItemIsEditable)
        if item_isWS:
            item.setText(0, item.text(0).replace(" *", ""))  # Remove visual cue
    elif column == 3 and (itemDict['Type'].startswith('Serie') or
                          itemDict['Type'] == "INTERPOLATION"): 
        item.setFlags(item.flags() | Qt.ItemIsEditable)
    elif column == 4 and itemDict['Type'].startswith('Serie'):
        item.setFlags(item.flags() | Qt.ItemIsEditable)
    else:
        item.setFlags(item.flags() & ~Qt.ItemIsEditable)

    tree_widget.blockSignals(False)

#========================================================================================
def show_dialog(title, fileHTML, width, height):
    with open(fileHTML, 'r') as file:
        html_text = file.read()
    
    dialog = QDialog()
    dialog.setWindowTitle(title)
    dialog.setFixedSize(width, height)
    
    main_layout = QVBoxLayout()
    dialog.setLayout(main_layout)
    text_browser = QTextBrowser()
    text_browser.setHtml(html_text)
    main_layout.addWidget(text_browser)
    
    button_layout = QHBoxLayout()
    button_layout.addItem(QSpacerItem(40, 20, QSizePolicy.Expanding, QSizePolicy.Minimum))
    ok_button = QPushButton('OK')
    ok_button.clicked.connect(dialog.accept)
    icon = QApplication.style().standardIcon(QStyle.SP_DialogApplyButton)
    ok_button.setIcon(icon)
    button_layout.addWidget(ok_button)
    main_layout.addLayout(button_layout)

    dialog.exec_()

#========================================================================================
def exit_confirm():

    reply = QMessageBox.question(
        main_window, 
        "Exit confirmation",
        "Are you sure you want to exit the application?",
        QMessageBox.Yes | QMessageBox.No,
        QMessageBox.No
    )
    
    if reply == QMessageBox.Yes:
        app.quit()

#========================================================================================
app = QApplication(sys.argv)

fontArial = QFont('Arial', 12)
app.setFont(fontArial)

icon = QIcon('resources/PyAnalySeries_icon.ico')
app.setWindowIcon(icon)

main_window = QMainWindow()
main_window.setWindowTitle("PyAnalySeries " + version)
main_window.setGeometry(100, 100, 1400, 600)

main_widget = QWidget()
layout = QVBoxLayout()

tree_widget = create_tree_widget()
tree_widget.setContextMenuPolicy(Qt.CustomContextMenu)
tree_widget.customContextMenuRequested.connect(show_context_menu)
tree_widget.itemChanged.connect(on_item_changed)
tree_widget.itemDoubleClicked.connect(on_item_double_clicked)

layout.addWidget(tree_widget)

main_widget.setLayout(layout)
main_window.setCentralWidget(main_widget)

menu_bar = main_window.menuBar()

#----------------------------------------------
file_menu = menu_bar.addMenu("File")

newWS_action = QAction("New worksheet", main_window)
newWS_action.setShortcut('Ctrl+n')
newWS_action.triggered.connect(new_WorkSheet)
openWS_action = QAction("Open worksheet", main_window)
openWS_action.setShortcut('Ctrl+o')
openWS_action.triggered.connect(open_WorkSheet)
saveWSs_action = QAction("Save worksheets", main_window)
saveWSs_action.setShortcut('Ctrl+s')
saveWSs_action.triggered.connect(save_WorkSheets)
import_action = QAction("Import data", main_window)
import_action.setShortcut('Ctrl+m')
import_action.triggered.connect(import_Series)
exit_action = QAction('Exit', main_window)
exit_action.setShortcut('q')
exit_action.triggered.connect(exit_confirm)

file_menu.addAction(newWS_action)
file_menu.addAction(openWS_action)
file_menu.addAction(saveWSs_action)
file_menu.addSeparator()
file_menu.addAction(import_action)
file_menu.addSeparator()
file_menu.addAction(exit_action)

#----------------------------------------------
edit_menu = menu_bar.addMenu("Edit")

cut_action = QAction("Cut", main_window)
cut_action.setShortcuts([QKeySequence("Ctrl+x"), QKeySequence(Qt.Key_Delete)])
cut_action.triggered.connect(cut_items)

copy_action = QAction("Copy", main_window)
copy_action.setShortcut(QKeySequence("Ctrl+c"))
copy_action.triggered.connect(copy_items)

paste_action = QAction("Paste", main_window)
paste_action.setShortcut(QKeySequence("Ctrl+v"))
paste_action.triggered.connect(paste_items)

display_action = QAction("Display Single", main_window)
display_action.setShortcut('Ctrl+d')
display_action.triggered.connect(displaySingleSerie_selected_series)

displayOverlaidSeries_action = QAction("Display Overlaid", main_window)
displayOverlaidSeries_action.setShortcut('Ctrl+t')
displayOverlaidSeries_action.triggered.connect(lambda: displayMultipleSeries_selected_series(overlaid=True))

displayStackedSeries_action = QAction("Display Stacked", main_window)
displayStackedSeries_action.setShortcut('Ctrl+r')
displayStackedSeries_action.triggered.connect(lambda: displayMultipleSeries_selected_series(overlaid=False))

close_all_action = QAction("Close all Display windows")
close_all_action.triggered.connect(close_all_windows)

edit_menu.addAction(cut_action)
edit_menu.addAction(copy_action)
edit_menu.addAction(paste_action)
edit_menu.addSeparator()
edit_menu.addAction(display_action)
edit_menu.addAction(displayOverlaidSeries_action)
edit_menu.addAction(displayStackedSeries_action)
edit_menu.addAction(close_all_action)

#----------------------------------------------
math_menu = menu_bar.addMenu("Processing")

defineFilter_action = QAction("Define Filter smoothing average", main_window)
defineFilter_action.setShortcut('Ctrl+f')
defineFilter_action.triggered.connect(define_filter)
applyFilter_action = QAction("Apply Filter smoothing average", main_window)
applyFilter_action.triggered.connect(apply_filter)

defineInterpolation_action = QAction("Define Interpolation", main_window)
defineInterpolation_action.setShortcut('Ctrl+i')
defineInterpolation_action.triggered.connect(define_interpolation)
applyInterpolationLinear_action = QAction("Apply Interpolation linear", main_window)
applyInterpolationLinear_action.triggered.connect(lambda: apply_interpolation('Linear'))
applyInterpolationPCHIP_action = QAction("Apply Interpolation PCHIP", main_window)
applyInterpolationPCHIP_action.setToolTip("This action applies PCHIP interpolation to the selected data.")
applyInterpolationPCHIP_action.triggered.connect(lambda: apply_interpolation('PCHIP'))

math_menu.addAction(defineFilter_action)
math_menu.addAction(applyFilter_action)
math_menu.addSeparator()
math_menu.addAction(defineInterpolation_action)
math_menu.addAction(applyInterpolationLinear_action)
math_menu.addAction(applyInterpolationPCHIP_action)

#----------------------------------------------
basicSeries_menu = menu_bar.addMenu("Basic series")

insolation_action = QAction("Insolation", main_window)
insolation_action.triggered.connect(define_insolationSerie)
astro_action = QAction("Astronomical parameters", main_window)
astro_action.triggered.connect(define_astroSerie)

basicSeries_menu.addAction(insolation_action)
basicSeries_menu.addSeparator()
basicSeries_menu.addAction(astro_action)

#----------------------------------------------
help_menu = menu_bar.addMenu('Help')

help_action = QAction('Help', main_window)
help_action.triggered.connect(lambda: show_dialog('About', 'resources/help.html', 1000, 800))
help_menu.addAction(help_action)

#----------------------------------------------
about_menu = menu_bar.addMenu('About')

about_action = QAction('About', main_window)
about_action.triggered.connect(lambda: show_dialog('About', 'resources/about.html', 1000, 600))
about_menu.addAction(about_action)

#----------------------------------------------
if filesWS:
    for fileWS in filesWS: 
        print('Loading...', fileWS)
        load_WorkSheet(fileWS)

#----------------------------------------------
main_window.setStatusBar(QStatusBar())
main_window.statusBar().showMessage('Application ready', 5000)
main_window.show()

#current_directory = os.getcwd()
#print(current_directory)

sys.exit(app.exec_())

